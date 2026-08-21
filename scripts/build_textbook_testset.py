#!/usr/bin/env python3
"""把「台灣手語教材」的句對整成獨立測試集。

來源：臺灣手語教材-搜尋網 https://tslcopus.deaf.com.tw/
      （教材影片來自臺灣手語教材資源網，教材編製 Jung-Hsing Chang；
        網頁空間由蒙恩聽障烘焙坊提供）
輸入：data/tsl_textbook/台灣手語教材.xlsx（欄位：原始中文／正確Gloss）

⚠️ **授權未查證**：該站首頁與列表頁**均無版權宣告或授權條款**。這與
`資料來源.md` 記錄的文化部語料庫／中正辭典不同——那兩者是 2026-08-04 明確
查證過「訓練＋散布皆合法，須標明出處」。沒有聲明**不等於**開放。故比照
`data/moe/` 的保守處理：**檔案留在 .gitignore，不推上公開 repo，論文附錄
不附原始句對**，報告中標明出處即可。要正式公開請先去信 tsl@deaf.com.tw。

⚠️ **不是教育部的資料**。交付時的檔名是「教育部手語資料」，但實際來源是上述
教材搜尋網，與 `data/moe/`（教育部常用手語辭典，7,913 詞條）是**完全不同的
來源與授權狀態**，不可混用或互相引用。

正規化（依據與 apply_corpus_review.py 相同）：既有 6,783 筆語料的 Gloss 中
全形標點出現 **0 次**，故清掉 `「」『』，。！？、；：`；ASCII 的 `?` `!` `()`
`+` 是既有記號予以保留。另清 token 內外的空白——本檔用 ` / ` 當分隔符裝飾，
既有語料則否（整批只有 5 個空格）。

用法：
    python3 scripts/build_textbook_testset.py            # 建表
    python3 scripts/build_textbook_testset.py --check    # 只報告不寫檔
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
SRC_DIR = BASE / "data" / "tsl_textbook"
XLSX = SRC_DIR / "台灣手語教材.xlsx"
OUT = SRC_DIR / "testset.jsonl"
SOURCE = ("臺灣手語教材-搜尋網 https://tslcopus.deaf.com.tw/"
          "（教材編製 Jung-Hsing Chang；授權未查證，不可公開原始句對）")

FULLWIDTH_PUNCT = re.compile(r"[「」『』，。！？、；：]")
# 訓練／既有語料，用來擋測試洩漏
CORPORA = [
    ("data/tslcorpus/parallel.jsonl", "文化部語料庫"),
    ("data/twtsl/twtsl_sentences.jsonl", "中正辭典例句"),
    ("data/synth/tsl_synth.jsonl", "規則模板合成句"),
    ("data/tsl_sentences.jsonl", "自有標記表"),
    ("data/papers/paper_examples_all.jsonl", "論文例句"),
]


def norm_gloss(text: str) -> str:
    stripped = FULLWIDTH_PUNCT.sub("", str(text))
    toks = []
    for t in stripped.split("/"):
        t = re.sub(r"\s+", "", t)      # token 內外空白都清（買+ + → 買++）
        if t:
            toks.append(t)
    return "/".join(toks)


def load_seen() -> dict[str, str]:
    seen = {}
    for rel, tag in CORPORA:
        p = BASE / rel
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            if line.strip():
                e = json.loads(line)
                zh = str(e.get("chinese", "")).strip()
                if zh:
                    seen.setdefault(zh, tag)
    return seen


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=XLSX)
    ap.add_argument("--check", action="store_true")
    args = ap.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"找不到 {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    raw = [(str(a).strip() if a else "", str(b).strip() if b else "")
           for a, b in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    raw = [r for r in raw if r[0] or r[1]]

    seen = load_seen()
    rows, stats, dropped = [], Counter(), []
    pairs_seen = set()
    for zh, gl in raw:
        stats["原始列"] += 1
        if not zh or not gl:
            stats["缺中文或Gloss"] += 1
            dropped.append((zh, "缺欄位"))
            continue
        g = norm_gloss(gl)
        if g != gl:
            stats["Gloss 經正規化"] += 1
        if not g:
            stats["正規化後為空"] += 1
            dropped.append((zh, "正規化後為空"))
            continue
        if (zh, g) in pairs_seen:
            stats["重複句對"] += 1
            dropped.append((zh, "與本檔內其他列重複"))
            continue
        # 測試洩漏：中文已存在於訓練用語料就不能當留存測試
        if zh in seen:
            stats["與既有語料重複（排除）"] += 1
            dropped.append((zh, f"已存在於{seen[zh]}"))
            continue
        pairs_seen.add((zh, g))
        rows.append({
            "id": f"TB{len(rows) + 1:04d}",
            "type": "sentence",
            "chinese": zh,
            "gloss": [t for t in g.split("/") if t],
            "gloss_text": g,
            "review_status": "source-provided",   # 來源標的「正確Gloss」，本團隊未複審
            "batch": "tsl-textbook-2026-08-22",
            "source": SOURCE,
        })

    print(f"輸出 {len(rows)} 句")
    for k, v in stats.most_common():
        print(f"  {k}: {v}")
    if dropped:
        print("  排除明細（前 10）:")
        for zh, why in dropped[:10]:
            print(f"    {zh[:28]}  ← {why}")
    lens = sorted(len(r["gloss"]) for r in rows)
    if lens:
        print(f"  Gloss 詞數：中位 {lens[len(lens)//2]}／平均 {sum(lens)/len(lens):.1f}"
              f"／最短 {lens[0]}／最長 {lens[-1]}")

    if args.check:
        print("--check：未寫檔")
        return 0
    SRC_DIR.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"寫出 {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

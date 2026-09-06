#!/usr/bin/env python3
"""產出兩個模型版本的盲測 A/B 評估表（例：v17cd vs v18）。

與 `make_human_eval_sheet.py` 的差別，以及為什麼需要另寫一支：

那一支比的是「模型輸出 vs 語料庫參考」，回答「模型的輸出可不可接受」。
本支比的是「模型 A vs 模型 B」，回答「哪一版比較好」——當自動指標互相
矛盾時（v18 教材集 BLEU +4.37 但 corpus −1.58／EM −2.41），只有人能仲裁。

**這個設計順帶修掉上一輪失敗的原因。** 2026-08-22 那份 v14 盲測回收後
判定無效，因為評分者是用「哪個符合參考資料」在打分——符合的三項全給 5、
另一版全給 1（見 [[tsl-blind-eval-form]]）。那等於繞過盲測。
模型對模型沒有這個漏洞：**兩欄都是模型輸出，沒有官方答案可對**，
評分者只能真的判斷手語品質。表中因此**完全不放參考 Gloss**。

其他防範：

- **只收兩版輸出不同的句子**。相同的送評是浪費，而且會讓評分者發現
  「有些題兩欄一樣」，進而猜測題目結構。
- **A/B 順序由句子雜湊決定**，可重現、與版本無關，評分者無法從位置推斷。
- **主判定是單選**（哪個較自然），不是 1–5 分。上一輪的三個 5 分量表正是
  被機械化填答的地方；單選比較難假填，也直接對應我們要的答案。
- 分項評分留著但標為選填，供想說明理由的評分者使用。

用法：
    python3 scripts/make_ab_eval_sheet.py --a v17cd --b v18script \
        --n-corpus 40 --n-textbook 40 --n-core 8

輸出：
    outputs/盲測_<A>_vs_<B>.xlsx        給評分者（不含答案）
    outputs/盲測_<A>_vs_<B>_key.json    對照 key（不要給評分者）
"""
from __future__ import annotations

import argparse
import hashlib
import json
import random
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(__file__).resolve().parent.parent
RESULTS = BASE / "results"
OUT = BASE / "outputs"
SPLITS = BASE / "data" / "splits"

SPLIT_LABEL = {"test": "核心短句", "test_corpus": "語料庫長句",
               "test_textbook": "教材句"}


def load_preds(tag: str, split: str) -> dict[str, list[str]]:
    p = RESULTS / f"{tag}_{split}.jsonl"
    if not p.exists():
        return {}
    out = {}
    for line in p.open(encoding="utf-8"):
        r = json.loads(line)
        try:
            ids = json.loads(r["raw"]).get("sign_ids") or []
        except Exception:                       # noqa: BLE001  破 JSON 視為空輸出
            ids = []
        out[r.get("id")] = [str(x) for x in ids]
    return out


def load_inventory() -> dict[str, str]:
    inv = {}
    for line in (BASE / "data" / "signs" / "sign_inventory.jsonl").open(encoding="utf-8"):
        o = json.loads(line)
        inv[o["sign_id"]] = o.get("gloss_clean") or o.get("gloss")
    return inv


def load_chinese(split: str) -> dict[str, str]:
    p = SPLITS / f"{split}.jsonl"
    if not p.exists():
        return {}
    return {json.loads(l)["id"]: json.loads(l)["chinese"]
            for l in p.open(encoding="utf-8") if l.strip()}


def to_gloss(ids: list[str], inv: dict[str, str]) -> str:
    return "/".join(inv.get(i, i) for i in ids)


def swap(sent_id: str) -> bool:
    """這題要不要左右對調。由句子 id 的雜湊決定——可重現，且與版本無關，
    所以評分者無法從「A 欄總是新版」這種規律反推。"""
    return int(hashlib.sha256(sent_id.encode()).hexdigest(), 16) % 2 == 1


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--a", required=True, help="模型 A 的 tag，例 v17cd")
    ap.add_argument("--b", required=True, help="模型 B 的 tag，例 v18script")
    ap.add_argument("--n-core", type=int, default=8)
    ap.add_argument("--n-corpus", type=int, default=40)
    ap.add_argument("--n-textbook", type=int, default=40)
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--exclude-key", type=Path, action="append", default=[],
                    help="先前輪次的 _key.json，其中的句子不再抽。擴大樣本時必用——"
                         "否則第二輪會與第一輪重疊，那是重問不是擴大。可重複給多份")
    ap.add_argument("--suffix", default="",
                    help="輸出檔名後綴（例 _round2）。擴大樣本時必用——第一輪的 xlsx "
                         "已由評分者填答，沒有後綴會被覆蓋掉")
    args = ap.parse_args()

    # 已用過的句子：擴大樣本時必須排除，而且要在抽樣前就排除，
    # 不然 rng.sample 抽到重複的再丟掉會讓各 split 的題數不足。
    used: set[str] = set()
    for kf in args.exclude_key:
        used |= {r["id"] for r in json.loads(kf.read_text(encoding="utf-8"))["rows"]}
    if used:
        print(f"排除先前 {len(args.exclude_key)} 輪共 {len(used)} 句", file=sys.stderr)

    inv = load_inventory()
    rng = random.Random(args.seed)
    picked = []
    for split, n in (("test", args.n_core), ("test_corpus", args.n_corpus),
                     ("test_textbook", args.n_textbook)):
        pa, pb = load_preds(args.a, split), load_preds(args.b, split)
        zh = load_chinese(split)
        # 只收兩版輸出不同、且**兩邊都非空**的。
        # 空輸出（模型吐出破 JSON 或空陣列）拿來問「哪個比較自然」沒有意義，
        # 評分者只能選另一邊，那一題就白問了。實測 437 題中有 1 題如此
        # （v18 在教材集的破 JSON 那列），單獨記錄比混進盲測有用。
        cand = sorted(i for i in (set(pa) & set(pb) & set(zh))
                      if pa[i] != pb[i] and pa[i] and pb[i] and i not in used)
        if not cand:
            print(f"⚠ {split}: 沒有兩版不同的句子", file=sys.stderr)
            continue
        if len(cand) < n:
            print(f"⚠ {split}: 只有 {len(cand)} 句可選，少於要求的 {n}", file=sys.stderr)
        for i in rng.sample(cand, min(n, len(cand))):
            picked.append({"id": i, "split": split, "chinese": zh[i],
                           "a_ids": pa[i], "b_ids": pb[i]})
    if not picked:
        print("沒有任何題目可出", file=sys.stderr)
        return 1
    rng.shuffle(picked)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "盲測"
    head = PatternFill("solid", fgColor="DDEBF7")
    bold = Font(bold=True)

    intro = [
        f"臺灣手語 Gloss 盲測評估表（{len(picked)} 題）",
        "",
        "請依「哪一版比較像自然的臺灣手語」作答。兩欄都是機器產生的候選，"
        "沒有哪一欄是標準答案，也沒有官方版本可以對照。",
        "",
        "填答方式：",
        "  1. 必填「較自然的版本」——從下拉選 A／B／兩者皆可／兩者皆不可。",
        "  2. 其餘欄位選填。若兩者皆不可，請盡量在「建議 Gloss」寫出您認為正確的版本。",
        "",
        "請不要依「看起來比較像課本／語料庫的寫法」來選，那不是本次要問的。"
        "請依實際打出來是否自然、意思是否傳達得到來判斷。",
    ]
    for r, t in enumerate(intro, 1):
        c = ws.cell(row=r, column=1, value=t)
        if r == 1:
            c.font = Font(bold=True, size=14)
    start = len(intro) + 2

    cols = ["編號", "類型", "中文", "版本 A", "版本 B", "較自然的版本",
            "A 語意(1-5)", "A 語序(1-5)", "B 語意(1-5)", "B 語序(1-5)",
            "建議 Gloss", "備註"]
    widths = [6, 12, 34, 34, 34, 15, 11, 11, 11, 11, 30, 24]
    for ci, (name, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=start, column=ci, value=name)
        c.font = bold
        c.fill = head
        ws.column_dimensions[c.column_letter].width = w

    key = []
    for n, item in enumerate(picked, 1):
        r = start + n
        sw = swap(item["id"])
        a_show = item["b_ids"] if sw else item["a_ids"]
        b_show = item["a_ids"] if sw else item["b_ids"]
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=SPLIT_LABEL.get(item["split"], item["split"]))
        for ci, val in ((3, item["chinese"]), (4, to_gloss(a_show, inv)),
                        (5, to_gloss(b_show, inv))):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        key.append({"no": n, "id": item["id"], "split": item["split"],
                    "chinese": item["chinese"],
                    "A": args.b if sw else args.a,
                    "B": args.a if sw else args.b,
                    "A_gloss": to_gloss(a_show, inv),
                    "B_gloss": to_gloss(b_show, inv)})

    last = start + len(picked)
    dv_pick = DataValidation(type="list",
                             formula1='"A,B,兩者皆可,兩者皆不可"', allow_blank=False)
    ws.add_data_validation(dv_pick)
    dv_pick.add(f"F{start + 1}:F{last}")
    dv5 = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(dv5)
    for col in ("G", "H", "I", "J"):
        dv5.add(f"{col}{start + 1}:{col}{last}")
    ws.freeze_panes = ws.cell(row=start + 1, column=1)

    OUT.mkdir(exist_ok=True)
    stem = f"盲測_{args.a}_vs_{args.b}{args.suffix}"
    if not args.suffix and (OUT / f"{stem}.xlsx").exists():
        # 沒給後綴又已有同名檔：多半是要擴大樣本卻忘了 --suffix，覆蓋掉的可能是
        # 評分者填好的表。寧可中止。
        print(f"✗ {OUT / f'{stem}.xlsx'} 已存在。擴大樣本請加 --suffix，"
              f"否則會覆蓋可能已填答的表。", file=sys.stderr)
        return 1
    xlsx = OUT / f"{stem}.xlsx"
    wb.save(xlsx)
    keyfile = OUT / f"{stem}_key.json"
    keyfile.write_text(json.dumps(
        {"note": "對照 key，**不要**給評分者。A/B 欄位記錄該題左右各是哪一版。",
         "a_tag": args.a, "b_tag": args.b, "seed": args.seed,
         "excluded_keys": [str(k) for k in args.exclude_key],
         "excluded_sentences": len(used),
         "n": len(picked), "rows": key}, ensure_ascii=False, indent=2),
        encoding="utf-8")

    from collections import Counter
    dist = Counter(SPLIT_LABEL.get(p["split"], p["split"]) for p in picked)
    print(f"寫出 {xlsx}（{len(picked)} 題）")
    print(f"  題型分佈：{dict(dist)}")
    print(f"  左右分佈：A 欄是 {args.a} 的有 "
          f"{sum(1 for k in key if k['A'] == args.a)} 題、"
          f"是 {args.b} 的有 {sum(1 for k in key if k['A'] == args.b)} 題")
    print(f"對照 key：{keyfile}（不要給評分者）")
    return 0


if __name__ == "__main__":
    sys.exit(main())

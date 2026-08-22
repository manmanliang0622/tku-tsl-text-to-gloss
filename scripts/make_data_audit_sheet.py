#!/usr/bin/env python3
"""產出「訓練資料 Gloss 抽查表」（教授回饋第 2 點：先把資料修好，再找熟悉臺灣手語的人抽查）。

**這張表和 make_human_eval_sheet.py 是兩件不同的事，不可互相取代：**

  make_human_eval_sheet.py  → 評「模型輸出」對不對，取樣自留存測試集，盲測 A/B
  本腳本                    → 審「訓練資料本身」對不對，取樣自 train，非盲測

教授第 2 點要的是後者：「有些論文裡的替代用法、錯誤標記、方向和空間資訊，可能
被直接串成同一個 Gloss 句子……應該先把資料解析修好，再找熟悉台灣手語的人幫忙
抽查，確認這些 Gloss 到底對不對。」解析已修（`clean_gloss_line`、`build_json_targets`
的疑問詞判定），但 train 的 5,347 句**從未經母語者逐句抽查**。

分層抽樣（每層都附「為什麼抽這層」，供老師理解取樣邏輯，不是隨機亂抽）：

  1. 語料庫·長度落差大   來源最大宗，且中文字數與 Gloss 詞數落差 >= 7，
                         是既有方法論標定的風險候選（見 資料來源.md 2026-07-22 第 6 點）
  2. 語料庫·一般         同來源對照組，用來估整體錯誤率而非只看最可疑的
  3. 辭典例句            中正辭典例句，多為短句，檢查是否為「中文照抄」
  4. 合成句·未審核       規則模板生成，train 內有 750 句 review_status=pending
  5. 合成句·已審核       2026-07-24 老師已審過的，當**信度對照組**——若老師這次
                         推翻自己審過的句子，代表抽查標準本身要重新對齊

輸出兩份：
  outputs/訓練資料抽查表_<split>.xlsx   交給老師的表
  outputs/訓練資料抽查_抽樣紀錄.json     抽樣參數與逐筆 id，供第三者重現／回填

回填：老師填完後，用「原始id」對回 data/splits/<split>.jsonl 與各來源檔。
本腳本**不會**改動任何資料，只讀取。

用法：
  python3 scripts/make_data_audit_sheet.py
  python3 scripts/make_data_audit_sheet.py --split train --n-corpus-risk 40 --seed 42
"""
import argparse
import json
import random
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(__file__).resolve().parent.parent
DATA = BASE / "data"
OUT = BASE / "outputs"

# 去標點後計中文字元數，與 資料來源.md 2026-07-22 第 6 點同一個門檻
PUNCT_RE = re.compile(r"[，。！？、；：「」『』（）()\[\],.!?;:\"'\s]")
RISK_DIFF = 7

PROBLEM_TYPES = ("選詞錯誤,語序錯誤,漏詞,多餘的詞,方向或空間資訊遺失,"
                 "照抄中文非手語詞,需看影片才能判斷,其他")


def gloss_len(gloss_text):
    return len([g for g in gloss_text.split("/") if g])


def length_gap(row):
    return abs(len(PUNCT_RE.sub("", row["chinese"])) - gloss_len(row["gloss_text"]))


def load_split(split):
    """讀切分並依 id 去重（train 經長度平衡過取樣，同一句會有多列）。"""
    path = DATA / "splits" / f"{split}.jsonl"
    if not path.exists():
        raise SystemExit(f"找不到 {path.relative_to(BASE)}，請先跑 scripts/split_data.py")
    rows = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            r = json.loads(line)
            rows.setdefault(r["id"], r)
    return list(rows.values())


def load_synth_status():
    """合成句的審核狀態，用來分「已審核／未審核」兩層。"""
    path = DATA / "synth" / "tsl_synth.jsonl"
    if not path.exists():
        return {}
    return {json.loads(l)["id"]: json.loads(l).get("review_status")
            for l in path.read_text(encoding="utf-8").splitlines() if l.strip()}


def build_strata(rows, synth_status):
    corpus = [r for r in rows if r["source"] == "tslcorpus"]
    strata = {
        "語料庫·長度落差大": dict(
            rows=[r for r in corpus if length_gap(r) >= RISK_DIFF],
            why=f"中文字數與 Gloss 詞數相差 >= {RISK_DIFF}，既有方法論標定的風險候選"),
        "語料庫·一般": dict(
            rows=[r for r in corpus if length_gap(r) < RISK_DIFF],
            why="同來源對照組，用來估整體錯誤率"),
        "辭典例句": dict(
            rows=[r for r in rows if r["source"] == "twtsl-sentence"],
            why="中正辭典例句，檢查是否為中文照抄"),
        "合成句·未審核": dict(
            rows=[r for r in rows if r["source"] == "synth"
                  and synth_status.get(r["id"]) == "pending"],
            why="規則模板生成，尚未經老師逐句審核"),
        "合成句·已審核": dict(
            rows=[r for r in rows if r["source"] == "synth"
                  and synth_status.get(r["id"]) not in (None, "pending")],
            why="2026-07-24 已審過，當信度對照組"),
    }
    return strata


def write_sheet(picked, path, split, seed):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "訓練資料抽查"

    intro = [
        "臺灣手語 訓練資料 Gloss 抽查表",
        "",
        "【這張表在問什麼】下面每一列都是**目前拿去訓練模型的資料**。",
        "請判斷「現行 Gloss」是不是該中文句子可接受的臺灣手語打法。",
        "這不是在評模型，是在檢查我們餵給模型的答案本身對不對。",
        "",
        "【請這樣填】",
        "  判定：正確／可接受但不道地／需修正／無法只憑文字判斷",
        "  若填「需修正」，請在【建議正確Gloss】寫下您認為正確的版本（用 / 分隔）",
        "  問題類型：可複選，用頓號分隔",
        "",
        "【說明】Gloss 用 / 分隔；「++」表重複、「＋」表複合詞，皆為原始標記，非錯字。",
        "「上下文」是同一段對話的前文，語料庫句子常需要前文才判斷得出來。",
        "非手部標記（表情／搖頭／揚眉）不在本次抽查範圍——那需要看影片，",
        "本專案的模型也不輸出這一層。若您覺得某句非看影片不可，請填「無法只憑文字判斷」。",
        "",
    ]
    for line in intro:
        ws.append([line])
    ws["A1"].font = Font(bold=True, size=14)

    header = ["編號", "抽樣層", "原始id", "中文句子", "現行Gloss", "上下文",
              "判定", "建議正確Gloss", "問題類型", "備註"]
    ws.append(header)
    hrow = ws.max_row
    for c in ws[hrow]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for i, r in enumerate(picked, 1):
        ws.append([i, r["_stratum"], r["id"], r["chinese"],
                   r["gloss_text"].replace("/", " / "),
                   (r.get("context") or "")[:120], "", "", "", ""])

    first, last = hrow + 1, ws.max_row
    dv_verdict = DataValidation(
        type="list",
        formula1='"正確,可接受但不道地,需修正,無法只憑文字判斷"', allow_blank=True)
    dv_problem = DataValidation(type="list", formula1=f'"{PROBLEM_TYPES}"',
                                allow_blank=True)
    ws.add_data_validation(dv_verdict)
    ws.add_data_validation(dv_problem)
    dv_verdict.add(f"G{first}:G{last}")
    dv_problem.add(f"I{first}:I{last}")

    for col, w in zip("ABCDEFGHIJ", [6, 18, 11, 38, 34, 34, 18, 30, 20, 20]):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=first, max_row=last):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = f"A{first}"

    meta = wb.create_sheet("抽樣說明")
    meta.append(["切分", split])
    meta.append(["亂數種子", seed])
    meta.append(["題數", len(picked)])
    meta.append([])
    meta.append(["抽樣層", "為什麼抽這層"])
    seen = {}
    for r in picked:
        seen.setdefault(r["_stratum"], r["_why"])
    for k, v in seen.items():
        meta.append([k, v])
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 70
    for row in meta.iter_rows():
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--split", default="train", help="要抽查的切分（預設 train）")
    ap.add_argument("--n-corpus-risk", type=int, default=40)
    ap.add_argument("--n-corpus-normal", type=int, default=40)
    ap.add_argument("--n-dict", type=int, default=25)
    ap.add_argument("--n-synth-pending", type=int, default=35)
    ap.add_argument("--n-synth-reviewed", type=int, default=10)
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    rows = load_split(args.split)
    strata = build_strata(rows, load_synth_status())
    quota = {
        "語料庫·長度落差大": args.n_corpus_risk,
        "語料庫·一般": args.n_corpus_normal,
        "辭典例句": args.n_dict,
        "合成句·未審核": args.n_synth_pending,
        "合成句·已審核": args.n_synth_reviewed,
    }

    rng = random.Random(args.seed)
    picked = []
    print(f"{args.split}：相異句 {len(rows)}")
    for name, want in quota.items():
        pool = strata[name]["rows"]
        take = sorted(pool, key=lambda r: r["id"])
        rng.shuffle(take)
        take = take[:want]
        if len(take) < want:
            print(f"  ⚠ {name}: 母體僅 {len(pool)} 句，要 {want} 句 → 全取 {len(take)}")
        else:
            print(f"  {name}: 母體 {len(pool)} 句 → 抽 {len(take)}")
        for r in take:
            r = dict(r, _stratum=name, _why=strata[name]["why"])
            picked.append(r)

    if not picked:
        raise SystemExit("沒有抽到任何句子。")

    rng.shuffle(picked)
    OUT.mkdir(exist_ok=True)

    path = OUT / f"訓練資料抽查表_{args.split}.xlsx"
    write_sheet(picked, path, args.split, args.seed)

    rec = OUT / "訓練資料抽查_抽樣紀錄.json"
    rec.write_text(json.dumps({
        "note": "抽樣紀錄，供第三者重現與回填；本腳本不改動任何資料。",
        "split": args.split, "seed": args.seed, "risk_diff_threshold": RISK_DIFF,
        "population": {k: len(v["rows"]) for k, v in strata.items()},
        "quota": quota,
        "rows": [{"編號": i, "抽樣層": r["_stratum"], "id": r["id"],
                  "source": r["source"], "group": r["group"],
                  "chinese": r["chinese"], "gloss_text": r["gloss_text"]}
                 for i, r in enumerate(picked, 1)],
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\nOK → {path.relative_to(BASE)}（{len(picked)} 題）")
    print(f"     {rec.relative_to(BASE)}")


if __name__ == "__main__":
    main()

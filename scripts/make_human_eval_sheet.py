#!/usr/bin/env python3
"""產出人工評估表（計畫 6.2；方法參考 CCL24-Eval 的 5 分制多維度評測）。

**設計核心：盲測 A/B**

本評估要回答的不是「模型幾分」，而是：
    當模型輸出與語料庫參考不同時，有多少其實也是可接受的臺灣手語？

理由：所有自動指標（EM/BLEU/ROUGE）都假設參考答案是唯一正解，但實測顯示
模型輸出常是「可理解的直譯」而非錯誤。例：
    中文：氣象預報說昨天晚上發布了海上颱風警報
    參考：氣象/宣布/昨天++/晚上/海/臺灣
    模型：天氣/預報/說/昨天/晚上/發布/海上颱風/警報
若這類輸出多屬可接受，則 EM 1.8% 嚴重低估真實可用性，我們也一直在錯誤的
訊號上調參數。

**為什麼要盲測**：若標明哪個是「語料庫官方答案」，評分者會有權威偏誤。
故 A/B 兩欄以句子雜湊決定順序（可重現），評分者不知孰為模型輸出。
對照答案 key 另存一檔，不放進給評分者的表內。

**取樣**：分層抽樣，涵蓋三個測試集；語料庫長句是目前最弱且最需判定的一群，
故佔比最高。

用法：
  舊格式（v11 等，結果檔已含 chinese/pred/ref）：
    python3 scripts/make_human_eval_sheet.py --model-tag v11 --n-corpus 60 --n-papers 30 --n-core 20
  腳本格式（v14 起，模型輸出是 sign_id）：
    python3 scripts/make_human_eval_sheet.py --model-tag v14script --format script \
        --n-corpus 50 --n-textbook 40 --n-core 10
"""
import argparse
import hashlib
import json
import random
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(__file__).resolve().parent.parent
RESULTS = BASE / "results"
OUT = BASE / "outputs"
SPLITS = BASE / "data" / "splits"

sys.path.insert(0, str(Path(__file__).resolve().parent))


def load_results(tag, split):
    """讀某模型在某測試集的逐句結果。"""
    for name in (f"{tag}_{split}_test.jsonl", f"{tag}_{split}.jsonl"):
        p = RESULTS / name
        if p.exists():
            return [json.loads(l) for l in p.read_text(encoding="utf-8").splitlines() if l.strip()]
    return []


def load_results_script(tag, split):
    """讀腳本格式（tsl-script-v1，v14 起）的逐句結果並還原成 gloss。

    與舊格式三個不同，都得在這裡處理：
      - 模型輸出是 sign_id 序列，須經 sign_inventory 轉回 gloss 才能給人評。
        總表查無的 ID（幻覺）去掉 TSL_ 前綴照列——若保留前綴，評分者一眼
        就看得出哪邊是機器，盲測就破了；哪些是幻覺記在對照 key。
      - 結果檔沒有中文與參考，須用 id 對回 data/splits/<split>.jsonl。
      - 參考一律用**完整參考**（原始 gloss_text）。`ref_sign_ids` 只含候選
        撈得到的詞，拿那個當參考等於把檢索缺口藏起來，會高估模型
        （同 eval_script_format.py 的 Full_reference 那組的理由）。
    """
    from eval_script_format import load_inventory, parse_output

    pred_path = RESULTS / f"{tag}_{split}.jsonl"
    split_path = SPLITS / f"{split}.jsonl"
    if not pred_path.exists() or not split_path.exists():
        return []
    id2gloss = load_inventory()
    meta = {}
    for line in split_path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            e = json.loads(line)
            meta.setdefault(e["id"], e)

    out = []
    for line in pred_path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        m = meta.get(r.get("id"))
        if not m:
            continue
        obj = parse_output(r.get("raw", "")) or {}
        pred_ids = [str(i) for i in obj.get("sign_ids", [])]
        pred = "/".join(id2gloss.get(i, i.removeprefix("TSL_")) for i in pred_ids)
        ref = "/".join(t.strip() for t in m["gloss_text"].split("/") if t.strip())
        out.append({
            "id": r["id"], "chinese": m["chinese"], "pred": pred, "ref": ref,
            "sign_ids": pred_ids,
            "hallucinated_ids": [i for i in pred_ids if i not in id2gloss],
            "needs_review": bool(obj.get("needs_review", False)),
            "oov_items": obj.get("oov_items", []),
        })
    return out


def side_for(chinese):
    """以句子雜湊決定模型輸出放 A 或 B（可重現、與抽樣無關）。"""
    h = hashlib.sha256(chinese.encode("utf-8")).hexdigest()
    return "A" if int(h[:8], 16) % 2 == 0 else "B"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--model-tag", default="v12", help="結果檔前綴，如 v11、v14script")
    ap.add_argument("--format", choices=("legacy", "script"), default="legacy",
                    help="legacy＝結果檔已含 chinese/pred/ref（v11 等）；"
                         "script＝tsl-script-v1（v14 起），sign_id 經總表還原、"
                         "參考用完整參考、教材句取代論文例句")
    ap.add_argument("--n-corpus", type=int, default=60)
    ap.add_argument("--n-papers", type=int, default=30, help="僅 legacy 格式使用")
    ap.add_argument("--n-textbook", type=int, default=40, help="僅 script 格式使用")
    ap.add_argument("--n-core", type=int, default=20)
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    rng = random.Random(args.seed)
    if args.format == "script":
        # 論文例句 2026-08-22 已停用，改教材句；核心 33 句的結果檔叫 <tag>_test
        sources = [
            ("語料庫長句", "test_corpus", args.n_corpus),
            ("教材句", "test_textbook", args.n_textbook),
            ("核心短句", "test", args.n_core),
        ]
        loader = load_results_script
    else:
        sources = [
            ("語料庫長句", "test_corpus", args.n_corpus),
            ("論文例句", "test_papers", args.n_papers),
            ("核心短句", "core33", args.n_core),
        ]
        loader = load_results

    rows = []
    for label, split, n in sources:
        recs = loader(args.model_tag, split)
        if not recs:
            print(f"  ⚠ 找不到 {args.model_tag}_{split} 的結果檔，略過「{label}」")
            continue
        # 只評「與參考不同」者才有判定價值；完全相同者另外計數即可。
        # 輸出為空（needs_review 拒答且無 sign_ids）者沒有 gloss 可評，也排除。
        diff = [r for r in recs if r.get("pred") and r["pred"] != r["ref"]]
        same = sum(1 for r in recs if r.get("pred") and r["pred"] == r["ref"])
        empty = len(recs) - len(diff) - same
        rng.shuffle(diff)
        picked = diff[:n]
        print(f"  {label}: 共 {len(recs)} 句（與參考相同 {same}、輸出為空 {empty}），"
              f"抽 {len(picked)} 句待判定")
        for r in picked:
            rows.append({"group": label, "chinese": r["chinese"],
                         "model": r["pred"], "ref": r["ref"],
                         **{k: r[k] for k in
                            ("id", "sign_ids", "hallucinated_ids",
                             "needs_review", "oov_items") if k in r}})

    if not rows:
        print("沒有可評估的資料，請先產生模型結果檔。")
        return

    rng.shuffle(rows)
    OUT.mkdir(exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人工評估"

    intro = [
        "臺灣手語翻譯 人工評估表（盲測）",
        "",
        "【請這樣填】每題有兩個手語 Gloss 版本（A 與 B），請分別評分，再回答哪個較自然。",
        "**兩個版本中有一個是電腦產生的，但不會告訴您是哪一個** —— 請純粹依手語本身判斷。",
        "",
        "評分維度（1–5 分，5 分最好）：",
        "  語意正確：是否完整、正確傳達中文句子的意思（有無漏掉或加入不該有的訊息）",
        "  語序自然：詞的順序是否符合臺灣手語的習慣，聾人看了會不會覺得怪",
        "  可否使用：若照這個順序打出來，聾人能不能看懂（就算不夠道地）",
        "",
        "「哪個較自然」欄請填：A／B／兩者皆可／兩者皆不可",
        "若兩個都不理想，請在【建議正確 Gloss】欄寫下您認為正確的版本（用 / 分隔）。",
        "",
    ]
    for line in intro:
        ws.append([line])
    ws["A1"].font = Font(bold=True, size=14)

    header = ["編號", "類型", "中文句子", "版本A", "版本B",
              "A語意正確", "A語序自然", "A可否使用",
              "B語意正確", "B語序自然", "B可否使用",
              "哪個較自然", "建議正確Gloss", "備註"]
    ws.append(header)
    hrow = ws.max_row
    for c in ws[hrow]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    key_rows = []
    for i, r in enumerate(rows, 1):
        model_side = side_for(r["chinese"])
        a = r["model"] if model_side == "A" else r["ref"]
        b = r["ref"] if model_side == "A" else r["model"]
        ws.append([i, r["group"], r["chinese"],
                   a.replace("/", " / "), b.replace("/", " / "),
                   "", "", "", "", "", "", "", "", ""])
        key_rows.append({"編號": i, "中文": r["chinese"], "模型輸出在": model_side,
                         "模型輸出": r["model"], "語料庫參考": r["ref"], "類型": r["group"],
                         **{k: r[k] for k in
                            ("id", "sign_ids", "hallucinated_ids",
                             "needs_review", "oov_items") if k in r}})

    first, last = hrow + 1, ws.max_row
    dv_score = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_pick = DataValidation(type="list",
                             formula1='"A,B,兩者皆可,兩者皆不可"', allow_blank=True)
    ws.add_data_validation(dv_score)
    ws.add_data_validation(dv_pick)
    for col in "FGHIJK":
        dv_score.add(f"{col}{first}:{col}{last}")
    dv_pick.add(f"L{first}:L{last}")

    for col, w in zip("ABCDEFGHIJKLMN",
                      [6, 12, 40, 34, 34, 9, 9, 9, 9, 9, 9, 13, 26, 18]):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=first, max_row=last):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = f"A{first}"

    path = OUT / f"人工評估表_{args.model_tag}_盲測.xlsx"
    wb.save(path)

    key_path = OUT / f"人工評估對照key_{args.model_tag}.json"
    key_path.write_text(json.dumps(
        {"note": "此檔為對照答案，**不可交給評分者**；回收後用於統計。",
         "model_tag": args.model_tag, "seed": args.seed, "rows": key_rows},
        ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\nOK → {path.relative_to(BASE)}（{len(rows)} 題）")
    print(f"     {key_path.relative_to(BASE)}  ⚠️ 對照答案，勿交給評分者")


if __name__ == "__main__":
    main()

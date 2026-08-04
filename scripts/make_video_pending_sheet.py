#!/usr/bin/env python3
"""產出「7 句待影片裁定」的人工裁定表（附文化部語料庫實證）。

背景：2026-07-24 手語老師審核時，967 句合成句中有 7 句無法只憑文字定案，
標記 teacher_train_eligible=False 排除於訓練。2026-08-04 以文化部語料庫
5,272 句真實聾人語料回查，其中 6 句的**文字層（Gloss token 序列）**已可據
實證裁定；NMS（疑問表情範圍、狀態變化）仍屬影片軌，不阻擋 gloss 層訓練。

輸出：data/synth/待影片裁定7句_人工裁定表.xlsx
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).resolve().parent.parent
OUT = BASE / "data" / "synth"

# 每句：現行 gloss、語料庫實證、建議、待影片項目
CASES = [
    dict(ids=["SYN0710"], chinese="你要去哪裡？", cur="你/去/哪裡/要",
         suggest="你/去/哪裡",
         evidence=[
             "語料庫「WH 疑問詞後接『要』」= 0 句（5,272 句中從未出現，現行寫法無實證）",
             "「要去哪裡申請？」→ 申請/哪裡（要、去 皆未打出）",
             "「刮除刀要去哪裡買？」→ 刮除刀/買/哪裡（要、去 皆未打出）",
             "「你們去哪兒玩？」→ 去/玩/哪裡（WH 句末）",
             "「哪裡」位於句末：21/29 句（72%）",
         ],
         reason="WH 與情態詞競爭句末位置時，語料顯示 WH 取得句末，『要』多半整個不打出。",
         video="疑問表情（揚眉／頭前傾）的起訖範圍"),
    dict(ids=["SYN0711", "SYN0712"], chinese="你今天要去哪裡？／今天你要去哪裡？",
         cur="今天/你/去/哪裡/要", suggest="今天/你/去/哪裡",
         evidence=["同上；另『時間詞置句首』已由語料實證 82%"],
         reason="同 SYN0710，去掉句末『要』；時間詞維持句首。",
         video="疑問表情範圍"),
    dict(ids=["SYN0713", "SYN0714"], chinese="你明天要去哪裡？／明天你要去哪裡？",
         cur="明天/你/去/哪裡/要", suggest="明天/你/去/哪裡",
         evidence=["同上"],
         reason="同 SYN0710。",
         video="疑問表情範圍"),
    dict(ids=["SYN0892"], chinese="你在哪裡上學？", cur="你/唸書/哪裡",
         suggest="你/讀書/哪裡",
         evidence=[
             "語料庫詞頻：讀書 61 句 vs 唸書 7 句（讀書為主流用詞）",
             "「你讀哪裡的啟聰學校？」→ 你/啟聰學校/讀書/哪裡",
             "「那你以前在哪裡讀碩士？」→ 你/以前/碩士/讀書/哪裡",
             "辭典兩詞皆收錄；『上學』辭典未收",
         ],
         reason="語序（WH 句末）本就正確，僅詞形建議由『唸書』改為語料主流的『讀書』。",
         video="兩詞若為同一手勢的異寫，則屬詞形選擇、無需影片"),
    dict(ids=["SYN0955"], chinese="我感冒了。", cur="我/感冒", suggest="我/感冒（維持不變）",
         evidence=[
             "「會淋濕感冒。」→ 淋雨/衣服/淋濕/感冒/會",
             "「…可能是感冒的關係。」→ 他/肺炎/支氣管炎/有/可能/感冒/影響",
             "語料庫中『感冒』皆為單一 gloss，未出現對應『了』的 token",
         ],
         reason="文字層已與語料一致（不打出『了』）。狀態變化屬 NMS，不影響 Text→Gloss 訓練標的。",
         video="狀態變化『了』的表情／動作幅度（屬影片軌，不阻擋文字層）"),
]


def build_vocab_rows():
    """比對 v_slots 詞形與語料庫真實用法，列出訊號矛盾的句子。"""
    corpus = [json.loads(l) for l in (BASE / "data/tslcorpus/parallel.jsonl")
              .read_text(encoding="utf-8").splitlines() if l.strip()]
    synth = [json.loads(l) for l in (BASE / "data/synth/tsl_synth.jsonl")
             .read_text(encoding="utf-8").splitlines() if l.strip()]
    twtsl = [json.loads(l) for l in (BASE / "data/twtsl/twtsl_words.jsonl")
             .read_text(encoding="utf-8").splitlines() if l.strip()]

    def cfreq(w):
        return sum(1 for r in corpus if w in r["gloss"])

    def status(w):
        if any(x.get("gloss_text") == w for x in twtsl):
            return "正式詞條"
        if any(w in (x.get("aliases") or []) for x in twtsl):
            return "同義索引名"
        return "未收錄"

    rows = []
    for used, alt in [("唸書", "讀書"), ("公共汽車", "公車"), ("疼", "痛")]:
        hits = [r for r in synth if used in r.get("gloss", [])]
        if not hits:
            continue
        ex = hits[0]
        rows.append([
            used, alt, f"{used} {cfreq(used)} vs {alt} {cfreq(alt)}",
            f"{used}={status(used)}／{alt}={status(alt)}",
            len(hits), "、".join(r["id"] for r in hits),
            f"{ex['chinese']} → {ex['gloss_text']}",
        ])
    return rows


VOCAB_ROWS = build_vocab_rows()


def main():
    OUT.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "7句裁定"

    ws.append(["「7 句待影片裁定」人工裁定表 — 附文化部語料庫（5,272 句真實聾人語料）實證"])
    ws.append(["填寫方式：看「語料庫實證」與「建議 Gloss」後，在【裁定】欄填 採納／修改／維持原樣／仍需影片；"
               "若填『修改』請在【裁定後 Gloss】寫下正確寫法"])
    ws.append([])
    header = ["代碼", "中文", "現行 Gloss", "建議 Gloss", "語料庫實證（可查證）",
              "建議理由", "仍待影片確認之項目", "【裁定】", "【裁定後 Gloss】", "【備註】"]
    ws.append(header)
    hrow = ws.max_row
    for c in ws[hrow]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")

    for case in CASES:
        ws.append([
            "、".join(case["ids"]), case["chinese"], case["cur"], case["suggest"],
            "\n".join("• " + e for e in case["evidence"]),
            case["reason"], case["video"], "", "", "",
        ])

    widths = [18, 26, 22, 22, 62, 40, 32, 14, 22, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=hrow + 1):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = f"A{hrow + 1}"

    # 第二分頁：原始證據句
    ws2 = wb.create_sheet("語料庫原始證據")
    ws2.append(["以下為文化部《臺灣手語語料庫》真實句，供裁定時直接比對"])
    ws2.append(["主題", "中文", "Gloss"])
    for c in ws2[2]:
        c.font = Font(bold=True)
    rows = [json.loads(l) for l in (BASE / "data/tslcorpus/parallel.jsonl")
            .read_text(encoding="utf-8").splitlines() if l.strip()]
    keys = ("哪裡", "唸書", "讀書", "感冒")
    for r in rows:
        if any(k in r["gloss"] for k in keys):
            ws2.append([r.get("theme", ""), r["chinese"], r["gloss_text"]])
    for i, w in enumerate([16, 46, 52], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    # 第三分頁：詞形一致性（v_slots 採辭典正式名，但語料庫偏好同義索引名）
    ws3 = wb.create_sheet("詞形一致性22句")
    ws3.append(["合成句採「辭典正式名」，但文化部語料庫（真實聾人用法）偏好另一詞形 —— "
                "兩者皆為同一辭典詞條（正式名／同義索引名），下游可播放性不受影響；"
                "問題在於訓練訊號互相矛盾（4,392 句語料教 A、22 句合成教 B）"])
    ws3.append(["合成句採用", "語料庫偏好", "語料庫詞頻", "辭典地位",
                "受影響句數", "受影響代碼", "例句", "【裁定】", "【裁定後詞形】", "【備註】"])
    for c in ws3[2]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    for row in VOCAB_ROWS:
        ws3.append(row + ["", "", ""])
    for i, w in enumerate([14, 14, 14, 26, 12, 46, 40, 14, 18, 20], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws3.iter_rows(min_row=3):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.freeze_panes = "A3"

    path = OUT / "待影片裁定7句_人工裁定表.xlsx"
    wb.save(path)
    print(f"OK → {path.relative_to(BASE)}")
    print(f"  裁定列 {len(CASES)} 組（涵蓋 7 句）／證據句 {ws2.max_row - 2} 句")


if __name__ == "__main__":
    main()

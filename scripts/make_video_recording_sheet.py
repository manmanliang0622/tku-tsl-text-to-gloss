#!/usr/bin/env python3
"""產出補片工作表：哪些詞該錄、哪些該先請顧問裁定。

`eval_video_coverage.py` 輸出的 data/video/video_gap.json 是逐詞缺口，
但那份不能直接拿去錄——裡面混了四種性質完全不同的東西：

  一致性動詞  「相信我」的「我」是方向變化，不是另一個手勢；詞根已在庫
  數量時間    「五萬」「六月」的組件都有，是組合規則的事
  量詞        缺的是「盤、顆、層、隻」這幾個字，不是那些詞
  類詞綴述語  「水倒入」「氣味飄散」是臨場構造的，本質上不能當詞條錄

真正該錄的是單音節動詞／名詞那一類。這支腳本把 gap 分流成
「該錄／該裁定／不必拍片」三張分頁，附上「現在虛擬人會演成什麼」，
讓演繹者與聾人顧問各自看得到自己要判斷的東西。

用法（先跑過 eval_video_coverage.py）：

    python scripts/make_video_recording_sheet.py

輸出：data/video/補片工作表.xlsx
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).resolve().parent.parent
GAP = BASE / "data" / "video" / "video_gap.json"
LEXICON = BASE / "data" / "video" / "lexicon.json"
OUT = BASE / "data" / "video" / "補片工作表.xlsx"

PRONOUNS = ("我", "你", "妳", "他", "她", "它")
NUM_CHARS = set("零一二三四五六七八九十百千萬億兩0123456789")


def greedy(word: str, keys: set[str]) -> list[str | None]:
    """composer.js tokenize 的貪婪切分；None 代表那個字沒有任何動作。"""
    out, i = [], 0
    while i < len(word):
        hit = next((word[i:i + n] for n in range(len(word) - i, 0, -1)
                    if word[i:i + n] in keys), None)
        out.append(hit)
        i += len(hit) if hit else 1
    return out


def classify(word: str, keys: set[str]) -> tuple[str, str]:
    """回傳 (分頁, 類別)。分頁決定這筆該給誰看。"""
    if re.fullmatch(r"[A-Za-z][A-Za-z\s]*", word):
        return "該錄", "外語（26 字母指拼可一次解決）"
    if len(word) == 1:
        return "該錄", "單音節動詞／名詞"
    for p in PRONOUNS:
        if (word.endswith(p) and word[:-1] in keys) or (word.startswith(p) and word[1:] in keys):
            return "不必拍片", "一致性動詞（我/你/他是方向，詞根已在庫）"
    head = "".join(c for c in word if c in NUM_CHARS)
    if head and word[0] in NUM_CHARS:
        unit = word[len(head):]
        if unit and unit not in keys:
            return "該錄", f"量詞／單位「{unit}」"
        if all(c in keys or c.isdigit() for c in head):
            return "不必拍片", "數量時間（組件已在庫，走組合規則）"
    pieces = greedy(word, keys)
    if None not in pieces and len(pieces) >= 2 and all(len(p) >= 2 for p in pieces):
        return "不必拍片", "複合詞，既有詞拼接可信"
    return "該裁定", "多字／拼接可疑（含類詞綴述語）"


SHEETS = {
    "該錄": ("這些是真的缺影片。同一個量詞（盤、顆、層…）錄一次就解掉整組，"
             "外語整批走 26 個字母指拼，詞條鍵用 fs:A…fs:Z。",
             ["詞", "出現次數", "類別", "現在會演成什麼", "【錄了嗎】", "【備註】"]),
    "該裁定": ("請聾人顧問判斷：既有詞的拼接算不算正確台灣手語？"
               "若是類詞綴述語（臨場構造，如「水倒入」「氣味飄散」）請標『不必收詞』。",
               ["詞", "出現次數", "現在會演成什麼", "【裁定】拼接可用／要錄／不必收詞",
                "【若要錄，正確詞形】", "【備註】"]),
    "不必拍片": ("這些不是缺影片，是查詢端或組合規則的事，列出來備查。",
                 ["詞", "出現次數", "類別", "現在會演成什麼"]),
}


def main() -> int:
    if not GAP.is_file() or not LEXICON.is_file():
        print(f"缺 {GAP.name} 或 {LEXICON.name}，先跑 scripts/eval_video_coverage.py")
        return 1
    keys = set(json.loads(LEXICON.read_text(encoding="utf-8")))
    gap = json.loads(GAP.read_text(encoding="utf-8"))

    rows: dict[str, list] = {k: [] for k in SHEETS}
    for e in gap:
        sheet, kind = classify(e["word"], keys)
        shown = "/".join(p or "·" for p in greedy(e["word"], keys)) or "（整詞不動）"
        rows[sheet].append((e["word"], e["n"], kind, shown))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, (note, header) in SHEETS.items():
        ws = wb.create_sheet(name)
        data = sorted(rows[name], key=lambda r: (-r[1], r[0]))
        ws.append([f"{name}：{len(data)} 詞／{sum(r[1] for r in data)} 次出現"])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([note])
        ws.append([])
        ws.append(header)
        for c in ws[4]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")
        for word, n, kind, shown in data:
            ws.append([word, n, kind, shown] if name != "該裁定" else [word, n, shown, "", "", ""])
        widths = [16, 10, 34, 30, 18, 22] if name != "該裁定" else [16, 10, 30, 30, 20, 22]
        for i, w in enumerate(widths[:len(header)], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=5):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A5"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK → {OUT.relative_to(BASE)}")
    for name in SHEETS:
        d = rows[name]
        print(f"  {name:<6} {len(d):>4} 詞 /{sum(r[1] for r in d):>5} 次")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

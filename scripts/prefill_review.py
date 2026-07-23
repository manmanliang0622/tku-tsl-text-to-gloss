#!/usr/bin/env python3
"""AI 語言學預審：填入合成句審核表的「審核結果／備註」欄（可重跑）。

立場（務必遵守，與 2026-07-23 全資料審核一致）：
  本腳本只做**可查證**的文字／語序層預審——依已查證學術文獻與 5,272 句真實
  語料的實證傾向（scripts/verify_rules_corpus.py）判定語序，並標出證據。
  凡需實際手形、移動方向、NMS 時間範圍、地區變體或影片語境者，一律標
  「需母語者裁定」，**不冒充母語聾人通過、不宣稱看過影片**。
  規則推導（rule-derived）句即使語序符合傾向，仍屬未經母語者裁定，一律標需覆核。

判定分級：
  - 「文字/語序層暫可」：僅限 attested-pattern（複製自有標記表實例）與
    corpus-attested（複製公開語料/辭典真實例句）——來源本身可靠。
  - 「語序符傾向，惟規則推導需覆核」：rule-derived 且語序符合語料庫主流傾向。
  - 「語序需母語者確認」：rule-derived 且語序不符主流傾向。
  - 「需母語者裁定」：T13/T14 之 WH「哪裡」與情態「要」句尾競合。
  以上再依句附加：NMS 需影片裁定、詞形需確認打法、坐/搭詞形需確認等。

用法：python3 scripts/prefill_review.py
輸出：就地更新 data/synth/review_sheet*.xlsx 的審核結果(空白列)＋備註欄，
      並在標題註明本表為 AI 語言學預審、須手語老師覆核。原資料 jsonl 不動。
"""
import json
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
DATA = BASE / "data"
SYNTH = DATA / "synth"
STAMP = "AI語言學預審 2026-07-23"

WORDFORM = {"公共汽車": "公車", "疼": "痛", "自行車": "騎腳踏車", "唸書": "讀書"}
NEG = {"不", "沒有", "沒", "不要", "不能", "不用", "別", "不行"}
WH = {"哪裡", "什麼", "幾點", "多少錢", "誰", "為什麼"}


def review(e):
    """回傳 (審核結果, 備註)。"""
    tid, conf, g = e["template_id"], e["confidence"], e["gloss"]
    ch = e["chinese"]
    notes = []

    if tid in ("T13", "T14"):
        v = "需母語者裁定：WH「哪裡」與情態「要」句尾競合"
        notes.append("Tai&Tsay2015 未給此競合逐句答案；情態句尾68%、WH句末63% 皆傾向非絕對")
    elif conf == "attested-pattern":
        v = "語序來源可靠（複製自有標記表實例），文字/語序層暫可"
    elif conf == "corpus-attested":
        v = "語序複製公開語料/辭典真實例句，文字/語序層暫可"
    else:  # rule-derived
        m = []
        negi = [i for i, t in enumerate(g) if t in NEG]
        if negi and negi[-1] >= len(g) - 2:
            m.append("否定句尾符語料庫76%")
        if g and g[0] in ("今天", "明天", "昨天", "現在"):
            m.append("時間句首符82%")
        if g and g[-1] == "要":
            m.append("情態句尾符68%(傾向)")
        if g and g[-1] in WH:
            m.append("WH句末符63%")
        if m:
            v = "語序符語料庫主流傾向，惟屬規則推導未經母語者裁定，需覆核"
            notes += m
        else:
            v = "語序需母語者確認（規則推導、語料無主流傾向可佐證）"

    if e.get("nms"):
        v += "；NMS需影片裁定"
        notes.append("NMS形式/時間範圍非文字可定（陳怡君2012：不同疑問功能表情不同）")
    wf = [t for t in g if t in WORDFORM]
    if wf:
        v += "；詞形需確認打法"
        for t in wf:
            notes.append(f"「{t}」為辭典主詞條(口語「{WORDFORM[t]}」)，辭典列同義索引，實際打法需母語者確認")
    if tid in ("T33", "T34"):
        notes.append("搭乘以Gloss『坐』表示，坐/搭/搭火車詞形需真實語料與影片確認(codex audit)")
    if "上班" in g and ("工作" in ch or "上班" in ch):
        notes.append("中文工作/上班對Gloss『上班』之語意(職業/活動/時段)需確認(codex audit)")

    note = f"[{STAMP}] " + "；".join(notes) if notes else f"[{STAMP}]"
    return v, note


def fill_sheet(path, synth):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = {c.value: c.column for c in ws[2] if c.value}
    c_id = header.get("編號")
    c_res = header.get("審核結果")
    c_note = header.get("備註")
    if not (c_id and c_res and c_note):
        print(f"  跳過（找不到欄）: {path.name}")
        return 0
    n = 0
    for r in range(3, ws.max_row + 1):
        sid = ws.cell(r, c_id).value
        e = synth.get(sid)
        if not e:
            continue
        v, note = review(e)
        ws.cell(r, c_res).value = v
        ws.cell(r, c_note).value = note
        n += 1
    # 標題加預審聲明
    t = ws.cell(1, 1)
    base = str(t.value or "").split("｜")[0]
    t.value = (base + "｜審核結果欄為 AI 語言學預審（依文獻＋語料實證）；"
               "標『需母語者裁定/確認』者仍須手語老師逐句覆核，非母語聾人最終判定")
    wb.save(path)
    return n


def main():
    synth = {json.loads(l)["id"]: json.loads(l)
             for l in (SYNTH / "tsl_synth.jsonl").read_text(encoding="utf-8").splitlines()}
    total = 0
    from collections import Counter
    dist = Counter()
    for e in synth.values():
        v, _ = review(e)
        if "需母語者裁定" in v:
            dist["需母語者裁定"] += 1
        elif "需母語者確認" in v.split("；")[0]:
            dist["語序需確認"] += 1
        elif "需覆核" in v.split("；")[0]:
            dist["語序符傾向惟需覆核"] += 1
        else:
            dist["文字/語序層暫可"] += 1
    for path in sorted(SYNTH.glob("review_sheet*.xlsx")):
        if path.name.startswith("~$"):
            continue
        n = fill_sheet(path, synth)
        total += n
        print(f"  已填 {n} 列 → {path.name}")
    print(f"OK: 預審 {total} 列")
    print("判定分布:", dict(dist))


if __name__ == "__main__":
    main()

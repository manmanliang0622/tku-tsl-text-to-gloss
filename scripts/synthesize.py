#!/usr/bin/env python3
"""規則模板資料合成（計畫 3.3 節步驟 A＋B）。

原則：
  1. 每個模板都以「自有標記表中實際存在的句型」＋「已查證的臺灣手語語法規則」為依據，
     規則出處見各模板的 rule_basis 欄位（對應計畫文件 3.3 節的 7 條規則）。
  2. 槽位只代入 Gloss 詞彙總表（data/tsl_gloss_vocab.json）內的詞彙，
     保證合成句的每個 Gloss 下游動作庫都有對應動作。
  3. 所有合成句 review_status=pending，須經手語老師／聾人顧問審核（計畫 3.3 節步驟 C）
     才能進入訓練集；同步產出 review_sheet.xlsx 審核表。
  4. confidence 分級：
     - attested-pattern：句型直接複製自有標記表已有句（僅替換地名槽位）
     - rule-derived   ：句型由查證過的語法規則組合推導，尚無自有影片實例，審核優先度最高

增量輸出（v2）：
  - 既有 tsl_synth.jsonl 中的句子保留原 SYN 編號不變，新句接續編號。
  - 審核表：首次執行寫 review_sheet.xlsx；之後每次只把「新增句」寫到
    review_sheet_batch{N}.xlsx，不覆蓋老師填寫中的舊審核表。

詞彙兩層制（v2）：
  - 第一層：自有詞彙表 data/tsl_gloss_vocab.json（每個 Gloss 都有自有動作影片）。
  - 第二層：中正大學手語辭典 data/twtsl/twtsl_words.jsonl（scrape_twtsl.py 產出，
    每詞附辭典示範影片網址）。2026-07-20 分工確認：語言模型端只負責翻譯語序，
    影片由下游端自行爬取；entry 的 external_glosses 欄標示「下游需另爬影片」的詞。

輸出：data/synth/tsl_synth.jsonl、data/synth/review_sheet*.xlsx
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

BASE = Path(__file__).resolve().parent.parent
DATA = BASE / "data"
OUT = DATA / "synth"

# ---------------------------------------------------------------- 詞彙槽位
# 縣市與離島（可自然接「住／上班／去／～人」）
CITIES = ["基隆", "台北", "新北", "桃園", "新竹", "苗栗", "台中", "南投",
          "彰化", "雲林", "嘉義", "台南", "高雄", "屏東", "宜蘭", "花蓮",
          "台東", "澎湖", "金門", "馬祖", "綠島", "蘭嶼"]
# 台北市內地區（住／上班／去 自然；「～人」不自然，故不用於身分句）
DISTRICTS = ["士林", "北投", "西門町", "中山", "淡水", "板橋", "中壢"]
PLACES = CITIES + DISTRICTS            # 29，住／上班／去
IDENTITY_PLACES = ["台灣"] + CITIES    # 23，「{P}人」身分句
TIMES = ["今天", "明天"]               # 時間詞（詞彙表內僅此二詞適用移動句）

# 外部詞彙槽位（中正大學手語辭典詞條，scrape_twtsl.py 產出後才可用；
# 這些詞有辭典示範影片、自有動作庫尚未收錄，合成句會標 external_glosses）
DRINKS = ["咖啡", "牛奶", "果汁", "紅茶", "珍珠奶茶"]
FOODS = ["水果", "蘋果", "香蕉", "麵包", "蛋糕", "早餐", "便當"]
DESTINATIONS = ["學校", "醫院", "公車站", "火車站"]  # 「車站」為火車站之同義索引名，不另立

# batch4 主題槽位（點餐／交通／問路／日常對話／看病；皆為辭典正式名，
# 語序證據見 data/refs/tslcorpus_evidence.jsonl 與 data/twtsl/twtsl_sentences.jsonl）
DRINKS2 = ["汽水", "豆漿", "咖啡", "牛奶", "果汁", "紅茶", "珍珠奶茶"]
FOODS2 = ["披薩", "漢堡", "三明治", "水餃", "麵", "火鍋", "蛋糕", "麵包"]
PRICE_ITEMS = ["咖啡", "牛奶", "果汁", "披薩", "漢堡", "三明治", "蛋糕", "麵包", "便當"]
LOCAL_PLACES = ["廁所", "學校", "醫院", "公車站", "火車站", "餐廳", "夜市", "圖書館", "公園"]
LOCAL_DEST = ["學校", "醫院", "夜市", "圖書館", "公園"]
CITY_DEST = ["台北", "台中", "台南", "高雄", "花蓮", "台東"]
BODY_PARTS = ["頭", "肚子", "牙齒", "腳", "手"]
OCCUPATIONS = ["老師", "學生", "醫生", "護士"]
# (中文說法, Gloss 正式名) 對，供 v_slots 用
VEHICLES_LOCAL = [("公車", "公共汽車"), ("計程車", "計程車")]
VEHICLES_CITY = [("火車", "火車"), ("高鐵", "高鐵")]

NMS_YESNO = "疑問表情（眉毛上揚、身體微前傾）貫穿全句，不比「嗎」"
NMS_WH = "疑問表情（眼睛、眉毛、頭部姿勢）搭配句末疑問詞"
NMS_NEG = "否定可伴隨搖頭、眼睛或嘴部表情"

# ---------------------------------------------------------------- 模板定義
# chinese: {P}=地點 {T}=時間；每個模板列出中文說法變體（同一 Gloss 目標）
TEMPLATES = [
    {"tid": "T1", "name": "定居句（住）",
     "chinese": ["我住在{P}。", "我住{P}。"],
     "gloss": ["我", "{P}", "住"], "nms": None, "slots": PLACES,
     "confidence": "attested-pattern",
     "rule_basis": "規則2 無情態詞之定居類動詞：[主語]/地點/動詞（張榮興2008 空間語法；"
                   "自有實例 P01 我住在台北→我/台北/住、S26 我住在＿→我/＿地點/住）"},
    {"tid": "T2", "name": "定居句（上班）",
     "chinese": ["我在{P}上班。", "我在{P}工作。"],
     "gloss": ["我", "{P}", "上班"], "nms": None, "slots": PLACES,
     "confidence": "attested-pattern",
     "rule_basis": "規則2 同 T1（自有實例 P02 我在新竹上班→我/新竹/上班）。"
                   "「工作」為中文端同義改寫，Gloss 沿用「上班」，需審核確認"},
    {"tid": "T3", "name": "是非問句（住）",
     "chinese": ["你住在{P}嗎？", "你是不是住在{P}？"],
     "gloss": ["你", "{P}", "住"], "nms": NMS_YESNO, "slots": PLACES,
     "confidence": "attested-pattern",
     "rule_basis": "規則2＋規則3 是非問句不比「嗎」、改用非手部標記貫穿全句"
                   "（Tai & Tsay 2015；自有實例 P03 你住在宜蘭嗎→你/宜蘭/住＋NMS）"},
    {"tid": "T4", "name": "時間＋情態移動句",
     "chinese": ["我{T}要去{P}。", "{T}我要去{P}。"],
     "gloss": ["{T}", "我", "去", "{P}", "要"], "nms": None,
     "slots": PLACES, "time_slots": TIMES,
     "confidence": "attested-pattern",
     "rule_basis": "規則1 情態詞「要」句尾＋規則5 時間詞句首（母語者例句「我要去面試→"
                   "我/去/工作/要」；自有實例 P04 我明天要去花蓮→明天/我/去/花蓮/要）"},
    {"tid": "T5", "name": "身分判斷句",
     "chinese": ["我是{P}人。"],
     "gloss": ["我", "{P}", "人"], "nms": None, "slots": IDENTITY_PLACES,
     "confidence": "attested-pattern",
     "rule_basis": "規則4 判斷句不比「是」：[主語]/[地點]/[身分]"
                   "（教育部課綱臺灣手語；自有實例 P05 我是桃園人→我/桃園/人）"},
    {"tid": "T6", "name": "情態移動句（無時間詞）",
     "chinese": ["我要去{P}。"],
     "gloss": ["我", "去", "{P}", "要"], "nms": None, "slots": PLACES,
     "confidence": "attested-pattern",
     "rule_basis": "規則1 情態詞「要」句尾（母語者例句「我要去買→我/去/買/要」；"
                   "T4 之無時間詞形，自有實例 S14 我要喝水→我/水/喝/要 同屬「要」句尾）"},
    {"tid": "T7", "name": "是非問句（身分）",
     "chinese": ["你是{P}人嗎？"],
     "gloss": ["你", "{P}", "人"], "nms": NMS_YESNO, "slots": IDENTITY_PLACES,
     "confidence": "rule-derived",
     "rule_basis": "規則4（P05 句型）＋規則3（P03 是非問句 NMS）之組合推導，"
                   "自有資料無此句型實例，審核優先"},
    {"tid": "T8", "name": "時間＋否定移動句",
     "chinese": ["我{T}不去{P}。", "{T}我不去{P}。"],
     "gloss": ["{T}", "我", "{P}", "去", "不"], "nms": NMS_NEG,
     "slots": PLACES, "time_slots": TIMES,
     "confidence": "rule-derived",
     "rule_basis": "規則6 否定詞置動詞後/句尾＋規則5 時間詞句首，直接類比 Jane Tsay 2021 "
                   "例句 TODAY-I-SCHOOL-GO-NOT（今天/我/學校/去/不），學校→地名；"
                   "無情態詞故地點在動詞前（v1.2 母語例句歸納）。自有資料無實例，審核優先"},
    {"tid": "T9", "name": "WH 疑問句（上班）",
     "chinese": ["你在哪裡上班？"],
     "gloss": ["你", "上班", "哪裡"], "nms": NMS_WH, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "規則7 WH 疑問詞置句末＋疑問 NMS（Tai & Tsay 2015），"
                   "類比自有實例 S25 你住哪裡→你/住/哪裡。自有資料無實例，審核優先"},
    # ------------------------------------------------------------ batch2 新句型
    {"tid": "T10", "name": "否定定居句",
     "chinese": ["我不住在{P}。", "我沒有住在{P}。"],
     "gloss": ["我", "{P}", "住", "不"], "nms": NMS_NEG, "slots": PLACES,
     "confidence": "rule-derived",
     "rule_basis": "規則6 否定詞置動詞後（自有實例 S18 我不舒服→我/舒服/不）"
                   "＋規則2 定居句語序（P01）。自有資料無此組合實例，審核優先"},
    {"tid": "T11", "name": "情態移動疑問句",
     "chinese": ["你要去{P}嗎？", "你是不是要去{P}？"],
     "gloss": ["你", "去", "{P}", "要"], "nms": NMS_YESNO, "slots": PLACES,
     "confidence": "rule-derived",
     "rule_basis": "規則1 情態詞「要」句尾（T6／S14）＋規則3 是非問句 NMS 不比「嗎」"
                   "（P03）。自有資料無此組合實例，審核優先"},
    {"tid": "T12", "name": "時間＋情態移動疑問句",
     "chinese": ["你{T}要去{P}嗎？", "{T}你要去{P}嗎？"],
     "gloss": ["{T}", "你", "去", "{P}", "要"], "nms": NMS_YESNO,
     "slots": PLACES, "time_slots": TIMES,
     "confidence": "rule-derived",
     "rule_basis": "規則1＋規則3＋規則5 時間詞句首之三規則組合（P04 語序＋P03 NMS），"
                   "審核優先"},
    {"tid": "T13", "name": "WH 移動疑問句",
     "chinese": ["你要去哪裡？"],
     "gloss": ["你", "去", "哪裡", "要"], "nms": NMS_WH, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "規則1 情態詞句尾＋規則7 WH 句末；「哪裡」與「要」的相對順序由"
                   "規則7（WH 句末）與規則1（情態句尾）競合，暫依 S14 語序推導"
                   "「要」最尾，務必請老師確認，審核最優先"},
    {"tid": "T14", "name": "時間＋WH 移動疑問句",
     "chinese": ["你{T}要去哪裡？", "{T}你要去哪裡？"],
     "gloss": ["{T}", "你", "去", "哪裡", "要"], "nms": NMS_WH,
     "slots": None, "time_slots": TIMES,
     "confidence": "rule-derived",
     "rule_basis": "T13＋規則5 時間詞句首。與 T13 同屬順序競合句型，審核最優先"},
    {"tid": "T15", "name": "否定移動句（無時間詞）",
     "chinese": ["我不去{P}。"],
     "gloss": ["我", "{P}", "去", "不"], "nms": NMS_NEG, "slots": PLACES,
     "confidence": "rule-derived",
     "rule_basis": "T8 之無時間詞形：規則6 否定詞置動詞後（Jane Tsay 2021 "
                   "TODAY-I-SCHOOL-GO-NOT 去掉時間詞），審核優先"},
    {"tid": "T16", "name": "是非問句（上班）",
     "chinese": ["你在{P}上班嗎？", "你是不是在{P}上班？"],
     "gloss": ["你", "{P}", "上班"], "nms": NMS_YESNO, "slots": PLACES,
     "confidence": "rule-derived",
     "rule_basis": "規則2 定居類語序（P02 上班）＋規則3 是非問句 NMS（P03），"
                   "審核優先"},
    {"tid": "T17", "name": "情態疑問句（喝水）",
     "chinese": ["你要喝水嗎？"],
     "gloss": ["你", "水", "喝", "要"], "nms": NMS_YESNO, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "自有實例 S14 我要喝水→我/水/喝/要 換主語＋規則3 是非問句 NMS，"
                   "審核優先"},
    {"tid": "T18", "name": "情態疑問句（廁所）",
     "chinese": ["你要去廁所嗎？"],
     "gloss": ["你", "廁所", "去", "要"], "nms": NMS_YESNO, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "自有實例 S15 我要上廁所→我/廁所/去/要 換主語＋規則3，審核優先"},
    {"tid": "T19", "name": "WH 價錢句",
     "chinese": ["這個多少錢？", "這多少錢？"],
     "gloss": ["這", "多少錢"], "nms": NMS_WH, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "自有實例 S30 我要這個→我/這/要（這）＋S27 多少錢＋規則7 WH 句末，"
                   "審核優先"},
    # ------------------------------------------ batch2 外部詞彙模板（twtsl 辭典詞）
    {"tid": "T20", "name": "情態飲食句（喝）",
     "chinese": ["我要喝{P}。"],
     "gloss": ["我", "{P}", "喝", "要"], "nms": None, "slots": DRINKS,
     "confidence": "rule-derived",
     "rule_basis": "自有實例 S14 我要喝水→我/水/喝/要 之飲料槽位替換（規則1 情態句尾）；"
                   "槽位詞取自中正手語辭典，尚無自有影片，審核優先"},
    {"tid": "T21", "name": "情態飲食疑問句（喝）",
     "chinese": ["你要喝{P}嗎？"],
     "gloss": ["你", "{P}", "喝", "要"], "nms": NMS_YESNO, "slots": DRINKS,
     "confidence": "rule-derived",
     "rule_basis": "T20 換主語＋規則3 是非問句 NMS（P03），審核優先"},
    {"tid": "T22", "name": "情態飲食句（吃）",
     "chinese": ["我要吃{P}。"],
     "gloss": ["我", "{P}", "吃", "要"], "nms": None, "slots": FOODS,
     "confidence": "rule-derived",
     "rule_basis": "S14 句型之動詞（喝→吃）與賓語槽位類推（規則1）；「吃」與槽位詞"
                   "均取自中正手語辭典，尚無自有影片，審核優先"},
    {"tid": "T23", "name": "情態飲食疑問句（吃）",
     "chinese": ["你要吃{P}嗎？"],
     "gloss": ["你", "{P}", "吃", "要"], "nms": NMS_YESNO, "slots": FOODS,
     "confidence": "rule-derived",
     "rule_basis": "T22 換主語＋規則3 是非問句 NMS（P03），審核優先"},
    {"tid": "T24", "name": "情態移動句（場所）",
     "chinese": ["我要去{P}。"],
     "gloss": ["我", "去", "{P}", "要"], "nms": None, "slots": DESTINATIONS,
     "confidence": "rule-derived",
     "rule_basis": "T6 之場所槽位擴充（規則1）；「學校」見 Jane Tsay 2021 例句，"
                   "槽位詞取自中正手語辭典，尚無自有影片，審核優先"},
    {"tid": "T25", "name": "時間＋否定移動句（場所）",
     "chinese": ["我{T}不去{P}。", "{T}我不去{P}。"],
     "gloss": ["{T}", "我", "{P}", "去", "不"], "nms": NMS_NEG,
     "slots": DESTINATIONS, "time_slots": TIMES,
     "confidence": "rule-derived",
     "rule_basis": "T8 之場所槽位版；「今天/我/學校/去/不」即 Jane Tsay 2021 "
                   "TODAY-I-SCHOOL-GO-NOT 原例句，其餘場所為類推，審核優先"},
    # ---------------------------------------- batch4 主題模板（語料庫查證後合成）
    # 證據來源：文化部臺灣手語語料庫（tslcorpus.moc.gov.tw，代碼＝篇章ID/句ID，
    # 摘錄存 data/refs/tslcorpus_evidence.jsonl）＋中正辭典例句（TWS 編號）
    {"tid": "T26", "name": "想望句（喝）〔點餐〕",
     "chinese": ["我想喝{P}。"],
     "gloss": ["{P}", "我", "想", "喝"], "nms": None, "slots": DRINKS2,
     "confidence": "corpus-attested",
     "rule_basis": "直接複製中正辭典例句「我想喝汽水…」→汽水/我/想/喝"
                   "（賓語主題化前置、「想」居動詞前），僅換飲料槽位"},
    {"tid": "T27", "name": "想望句（吃）〔點餐〕",
     "chinese": ["我想吃{P}。"],
     "gloss": ["{P}", "我", "想", "吃"], "nms": None, "slots": FOODS2,
     "confidence": "rule-derived",
     "rule_basis": "T26 換動詞；賓語前置 O-V 另見文化部語料庫 G1D16P1/a102 "
                   "我/素/吃、G1D15P1/a73 早餐/吃++，審核優先"},
    {"tid": "T28", "name": "價錢句（品項）〔點餐〕",
     "chinese": ["{P}多少錢？"],
     "gloss": ["{P}", "多少錢"], "nms": NMS_WH, "slots": PRICE_ITEMS,
     "confidence": "rule-derived",
     "rule_basis": "T19＋主題+WH 句末結構（文化部語料庫 G3C12/a94 百貨公司在哪裡"
                   "→百貨公司/哪裡 同構；規則7），審核優先"},
    {"tid": "T29", "name": "好吃疑問句〔點餐〕",
     "chinese": ["{P}好吃嗎？"],
     "gloss": ["{P}", "好吃"], "nms": NMS_YESNO, "slots": FOODS2,
     "confidence": "rule-derived",
     "rule_basis": "規則4 判斷句無繫詞＋形容詞句尾（文化部語料庫 G1C32/a7 "
                   "吃/健康/對/身體/好、G1D3P1/a72 選擇/火車/好）＋規則3 NMS，審核優先"},
    {"tid": "T30", "name": "場所 WH 句〔問路〕",
     "chinese": ["{P}在哪裡？"],
     "gloss": ["{P}", "哪裡"], "nms": NMS_WH, "slots": LOCAL_PLACES,
     "confidence": "corpus-attested",
     "rule_basis": "文化部語料庫 G3C12/a94 百貨公司在哪裡？→百貨公司/哪裡 直接同構"
                   "（主題＋WH 句末，規則7），僅換場所槽位；廁所版另見 T18 註"},
    {"tid": "T31", "name": "讀書 WH 句〔問路〕",
     "chinese": ["你在哪裡讀書？", "你在哪裡上學？"],
     "gloss": ["你", "唸書", "哪裡"], "nms": NMS_WH, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "類比文化部語料庫 G2C31/a162 你哪裡畢業？→你/畢業/哪裡＋T9 句構；"
                   "「唸書」＝辭典正式名（讀書＝其同義索引名），審核優先"},
    {"tid": "T32", "name": "腳踏車移動句〔交通〕",
     "chinese": ["我每天騎腳踏車去{P}。"],
     "gloss": ["我", "每天", "自行車", "去", "{P}"], "nms": None,
     "slots": LOCAL_DEST,
     "confidence": "corpus-attested",
     "rule_basis": "直接複製中正辭典例句 TWS0047 我每天騎腳踏車上學→"
                   "我/每天/騎腳踏車/去/讀書，換目的地槽位；"
                   "「自行車」＝騎腳踏車之辭典正式名，請審核確認詞形"},
    {"tid": "T33", "name": "市內搭乘句〔交通〕",
     "chinese": ["我要坐{V}去{P}。", "我要搭{V}去{P}。"],
     "gloss": ["我", "{V}", "坐", "去", "{P}", "要"], "nms": None,
     "slots": LOCAL_DEST, "v_slots": VEHICLES_LOCAL,
     "confidence": "rule-derived",
     "rule_basis": "文化部語料庫 G3C39/a143 你每天都坐這班公車上學嗎？→"
                   "你/每天/公車/公車+這/坐車/去/讀書 之語序＋規則1 情態「要」句尾；"
                   "「公共汽車」＝公車正式名，審核優先"},
    {"tid": "T34", "name": "城際搭乘句〔交通〕",
     "chinese": ["我要坐{V}去{P}。", "我要搭{V}去{P}。"],
     "gloss": ["我", "{V}", "坐", "去", "{P}", "要"], "nms": None,
     "slots": CITY_DEST, "v_slots": VEHICLES_CITY,
     "confidence": "rule-derived",
     "rule_basis": "同 T33 句構；城際版另證：文化部語料庫 G2D13P1/a163 "
                   "我以前搭火車去台北唸書→像/我/以前/搭火車/去/台北/唸書，審核優先"},
    {"tid": "T35", "name": "部位疼痛句〔看病〕",
     "chinese": ["我{P}痛。", "我{P}很痛。"],
     "gloss": ["我", "{P}", "疼"], "nms": None, "slots": BODY_PARTS,
     "confidence": "rule-derived",
     "rule_basis": "文化部語料庫 G1D13P1/a127 …我/右肩/抬起/痛/會（部位＋痛）；"
                   "「疼」＝辭典正式名（痛＝其同義索引名），請審核確認詞形"},
    {"tid": "T36", "name": "吃藥句〔看病〕",
     "chinese": ["我要吃藥。"],
     "gloss": ["我", "藥", "吃", "要"], "nms": None, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "中正辭典 TWS0132 咳嗽要記得吃藥→咳嗽/藥/吃/記得（藥/吃 O-V）"
                   "＋規則1 情態句尾（S14 我/水/喝/要 同構），審核優先"},
    {"tid": "T37", "name": "吃藥疑問句〔看病〕",
     "chinese": ["你要吃藥嗎？"],
     "gloss": ["你", "藥", "吃", "要"], "nms": NMS_YESNO, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "T36 換主語＋規則3 是非問句 NMS（P03），審核優先"},
    {"tid": "T38", "name": "看醫生疑問句〔看病〕",
     "chinese": ["你要看醫生嗎？"],
     "gloss": ["你", "看醫生", "要"], "nms": NMS_YESNO, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "自有實例 S20 我要看醫生→我/看醫生/要 換主語＋規則3，審核優先"},
    {"tid": "T39", "name": "感冒句〔看病〕",
     "chinese": ["我感冒了。", "我感冒。"],
     "gloss": ["我", "感冒"], "nms": None, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "狀態句無繫詞（規則4）；中正辭典 TWS0437 妹妹生病了→妹妹/生病 "
                   "同構，審核優先"},
    {"tid": "T40", "name": "感冒疑問句〔看病〕",
     "chinese": ["你感冒了嗎？"],
     "gloss": ["你", "感冒"], "nms": NMS_YESNO, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "T39 換主語＋規則3 是非問句 NMS（P03），審核優先"},
    {"tid": "T41", "name": "天氣句〔日常對話〕",
     "chinese": ["今天天氣很好。"],
     "gloss": ["今天", "天氣", "好"], "nms": None, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "規則5 時間詞句首＋形容詞句尾（文化部語料庫 G1C32/a7 "
                   "…對/身體/好），審核優先"},
    {"tid": "T42", "name": "名字 WH 句（第三人稱）〔日常對話〕",
     "chinese": ["他叫什麼名字？"],
     "gloss": ["他", "名字", "什麼"], "nms": NMS_WH, "slots": None,
     "confidence": "rule-derived",
     "rule_basis": "自有實例 S23 你叫什麼名字→你/名字/什麼 換主語＋規則7，審核優先"},
    {"tid": "T43", "name": "職業判斷句〔日常對話〕",
     "chinese": ["我是{P}。"],
     "gloss": ["我", "{P}"], "nms": None, "slots": OCCUPATIONS,
     "confidence": "rule-derived",
     "rule_basis": "規則4 判斷句不比「是」（P05 我是桃園人→我/桃園/人 同構），"
                   "審核優先"},
    {"tid": "T44", "name": "職業是非問句〔日常對話〕",
     "chinese": ["你是{P}嗎？"],
     "gloss": ["你", "{P}"], "nms": NMS_YESNO, "slots": OCCUPATIONS,
     "confidence": "rule-derived",
     "rule_basis": "T43 換主語＋規則3 是非問句 NMS（T7 同構），審核優先"},
]


def load_existing():
    """讀既有真實資料：中文句集合（避免合成句與原句重複）＋兩層合法 Gloss 集合。

    回傳 (中文句集合, 自有詞彙集合, 外部詞彙集合)。
    外部詞彙＝中正大學手語辭典詞條（scrape_twtsl.py 產出），有辭典示範影片
    但自有動作庫尚未收錄；檔案不存在時為空集合。
    """
    chinese = set()
    for line in (DATA / "tsl_sentences.jsonl").read_text(encoding="utf-8").splitlines():
        chinese.add(json.loads(line)["chinese"])
    vocab = set(json.load((DATA / "tsl_gloss_vocab.json").open(encoding="utf-8"))["glosses"])
    external = set()
    twtsl = DATA / "twtsl" / "twtsl_words.jsonl"
    if twtsl.exists():
        for line in twtsl.read_text(encoding="utf-8").splitlines():
            external.add(json.loads(line)["gloss_text"])
    return chinese, vocab, external


def expand():
    existing_chinese, vocab, external = load_existing()
    entries, seen = [], set()
    for t in TEMPLATES:
        slot_vals = t["slots"] or [None]
        time_vals = t.get("time_slots") or [None]
        # v_slots：中文說法與 Gloss 正式名不同的槽位，元素為 (中文, Gloss) 對
        # （如 ("公車", "公共汽車")；Gloss 一律用辭典正式名供下游檢索）
        v_vals = t.get("v_slots") or [None]
        for p in slot_vals:
          for v in v_vals:
            v_zh, v_gl = v if v else ("", "")
            for tm in time_vals:
                gloss = [g.replace("{P}", p or "").replace("{T}", tm or "")
                         .replace("{V}", v_gl)
                         for g in t["gloss"]]
                unknown = [g for g in gloss if g not in vocab and g not in external]
                if unknown:
                    hint = ("（外部詞彙檔 data/twtsl/twtsl_words.jsonl 不存在，"
                            "請先執行 scripts/scrape_twtsl.py）" if not external else "")
                    raise ValueError(f"{t['tid']} 產生了詞彙表外 Gloss: {unknown}{hint}")
                ext_used = [g for g in gloss if g not in vocab]
                for vi, ch_tpl in enumerate(t["chinese"]):
                    ch = (ch_tpl.replace("{P}", p or "").replace("{T}", tm or "")
                          .replace("{V}", v_zh))
                    if ch in existing_chinese:   # 與真實資料重複（如 T1×台北=P01）
                        continue
                    key = (ch, "/".join(gloss))
                    if key in seen:
                        continue
                    seen.add(key)
                    entries.append({
                        "id": f"SYN{len(entries)+1:04d}",
                        "type": "sentence",
                        "chinese": ch,
                        "gloss": gloss,
                        "gloss_text": "/".join(gloss),
                        "nms": t["nms"],
                        "template_id": t["tid"],
                        "template_name": t["name"],
                        "is_paraphrase": vi > 0,
                        "confidence": t["confidence"],
                        "rule_basis": t["rule_basis"],
                        "external_glosses": ext_used,
                        "review_status": "pending",
                        "batch": "規則模板合成",
                    })
    return entries


def write_review_sheet(entries, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "審核表"
    ws.append(["合成句審核表 — 請手語老師／聾人顧問逐句確認 Gloss 與 NMS "
               "（審核結果填：通過／修正／不通過；rule-derived 句型請優先審核）"])
    header = ["編號", "模板", "信心等級", "中文", "TSL Gloss", "NMS",
              "規則依據", "外部詞彙（尚無自有影片）", "審核結果", "修正後 Gloss", "備註"]
    ws.append(header)
    for c in ws[2]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    for e in entries:
        ws.append([e["id"], f"{e['template_id']} {e['template_name']}",
                   e["confidence"], e["chinese"], e["gloss_text"],
                   e["nms"] or "", e["rule_basis"],
                   "、".join(e.get("external_glosses") or []), "", "", ""])
    widths = [10, 22, 16, 26, 26, 30, 60, 22, 10, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    wb.save(path)


def main():
    """增量輸出：既有句保留原 SYN 編號與內容，新句接續編號並另出批次審核表。"""
    OUT.mkdir(exist_ok=True)
    entries = expand()
    jsonl_path = OUT / "tsl_synth.jsonl"

    old = []
    if jsonl_path.exists():
        old = [json.loads(l) for l in
               jsonl_path.read_text(encoding="utf-8").splitlines()]
    old_keys = {(e["chinese"], e["gloss_text"]) for e in old}
    new_entries = [e for e in entries
                   if (e["chinese"], e["gloss_text"]) not in old_keys]

    sheets = [p for p in OUT.glob("review_sheet*.xlsx")
              if not p.name.startswith("~$")]
    batch_no = len(sheets) + 1
    next_num = max((int(e["id"][3:]) for e in old), default=0) + 1
    for e in new_entries:
        e["id"] = f"SYN{next_num:04d}"
        next_num += 1
        if old:
            e["batch"] = f"規則模板合成-batch{batch_no}"

    all_entries = old + new_entries
    with jsonl_path.open("w", encoding="utf-8") as f:
        for e in all_entries:
            f.write(json.dumps(e, ensure_ascii=False) + "\n")
    sheet_path = OUT / ("review_sheet.xlsx" if not old
                        else f"review_sheet_batch{batch_no}.xlsx")
    write_review_sheet(new_entries if old else all_entries, sheet_path)

    from collections import Counter
    by_t = Counter(e["template_id"] for e in new_entries)
    by_c = Counter(e["confidence"] for e in new_entries)
    ext = sum(1 for e in new_entries if e.get("external_glosses"))
    print(f"OK: 總計 {len(all_entries)} 句（既有 {len(old)}＋新增 {len(new_entries)}）"
          f"→ {jsonl_path.relative_to(BASE)}")
    print("新增各模板:", dict(sorted(by_t.items())))
    print("新增信心分布:", dict(by_c))
    if ext:
        print(f"含外部詞彙句數: {ext}（Gloss 尚無自有動作影片，見審核表欄位）")
    print(f"本批審核表: {sheet_path.relative_to(BASE)}")


if __name__ == "__main__":
    main()
